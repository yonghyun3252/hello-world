from openpyxl import Workbook
from random import *
wb = Workbook()
ws=wb.active


#1줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"]) #성적표
for i in range(1,11):
    ws.append([i, randint(0,100), randint(0,100)])

col_B = ws["B"]
# print(col_B)
# for cell in col_B:
#     print(cell.value)

col_range = ws["B:C"]
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

row_title = ws[1] #1번째 row만 가지고 오기
# for cell in row_title:
#     print(cell.value)


row_range = ws[2:6]
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end = " ")
#     print()

# from openpyxl.utils.cell import coordinate_from_string

# row_range = ws[2:ws.max_row] #2번째 줄부터 마지막까지
# for rows in row_range:
#     for cell in rows:
#         # print(cell.value, end=" ")
#         # print(cell.coordinate, end=" ")
#         xy = coordinate_from_string(cell.coordinate)
#         # print(xy, end=" ")
#         print(xy[0], end="")
#         print(xy[1], end=" ")
#     print()



#전체 rows
# print(tuple(ws.rows))
# for row in tuple(ws.rows):
#     print(row[0].value) #

# #전체 columns
# print(tuple(ws.columns))

# for column in tuple(ws.columns):
#     print(column[10].value)

# for row in ws.iter_rows():
#     print(row[1].value)

for column in ws.iter_cols():
    print(column[0].value)
wb.save("sample.xlsx")

