from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "yonghyunSheet"

ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"]) #A1 셀의 정보를 출력

print(ws["A1"].value) #A1 셀의 값을 출력
print(ws["A10"].value) # 값이 없을때는 none 출력
print(ws.cell(row=1, column=1).value) # row = 1,2,3.... column = A(1), B(2), C(3)
print(ws.cell(row =1, column = 2).value)


c = ws.cell(row = 1, column = 3, value=10) #ws["C1"].value = 10
print(c.value)


from random import *
#반복문
index = 1
for x in range(1,11):
    for y in range(1,11):
        # ws.cell(row =x, column=y,value=randint(0,100)) #0~100사이의 숫자를 row 10개 column 10개채움
        ws.cell(row=x, column=y, value=index)
        index +=1
wb.save("sample.xlsx")
