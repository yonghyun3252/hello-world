from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() #새로운 sheet 기본 이름으로 생성


ws.title = "MySheet" #sheet 이름 변경
ws.sheet_properties.tabColor = "ff66ff" # RGB 형태로 값을 넣어주면 탭 색상 변경

#sheet, Mysheet, Yoursheet
ws1 = wb.create_sheet("YourSheet") #주어진이름으로 sheet생성
ws2 = wb.create_sheet("NewSheet", 2) # 2번째 index에 sheet 생성
ws2.sheet_properties.tabColor = "ff66fa"
wb.save("sample.xlsx")

new_ws = wb["NewSheet"] # Dict 형태로 sheet에 접근

print(wb.sheetnames) # 모든 sheet 이름 확인

#sheet 복사
new_ws["A1"] = "Test" #A1 란에 Test라는 글자 입력
target = wb.copy_worksheet(new_ws) # target이라는 변수안에 new_ws값에 저장되어 있는 NewSheet 복사
target.title = "Copied Sheet" # target의 복사된 sheet 이름을 Copied

wb.save("sample.xlsx")