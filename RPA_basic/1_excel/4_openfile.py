from openpyxl import load_workbook
wb = load_workbook("sample.xlsx") #파일에서 워크북을 불러옴
ws = wb.active #활성화된 sheet

#cell 데이터 불러오기
for x in range(1,11):
    for y in range(1,11):
        print(ws.cell(row=x, column=y).value, end=" ")
    
    print()


# cell 갯수를 모를때
for x in range(1, ws.max_row+1):
    for y in range(1, ws.max_column+1):
        print(ws.cell(row=x, column=y).value, end=" ")
    
    print()